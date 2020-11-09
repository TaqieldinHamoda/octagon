import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# Global Variables and references to data. DO NOT MODIFY
EXCEL = "./Data.xlsx"
SEX = ('ALL', 'F', 'M')
PROV = ('ALL', 'AB', 'ON', 'QC', 'BC', 'Atlantic', 'Prairies', 'UNKWN')
CON_ACT = ('ALL', 'No', 'Yes', 'Null')
MEASURE = ('Tx', 'events', 'censored')
AGES = tuple(["ALL", "18-19"] + [f"{i}-{i+4}" for i in range(20, 61, 5)] + ["65+"])


def get_data():
    """
    :return: A pandas data frame with all of the data
    """
    turn_int = lambda x: 0 if x is None else int(x)
    sheet = load_workbook(filename=EXCEL, read_only=True, data_only=True).active
    legend = []
    data = []

    for row in sheet.iter_rows(values_only=True):
        legend.append(row[:5])
        data.append(tuple(map(turn_int, row[5:])))

    legend = pd.DataFrame(np.array(legend), columns=("prov", "con_act", "sex", "age", "measure"))
    data = pd.DataFrame(np.array(data), columns=[f"M{i}" for i in range(40)])

    return pd.concat([legend, data], axis=1, sort=False)


def filter_data(prov=PROV, con_act=CON_ACT, sex=SEX, ages=AGES, measure=MEASURE, m_start=0, m_end=39):
    """
    :param prov: A list with all provinces that you want to get data for
    :param con_act: A list with all cancer statuses that you want to get data for
    :param sex: A list with all sex that you want to get data for
    :param ages: A list with all ages that you want to get data for
    :param measure: A list with all measures that you want to get data for
    :param m_start: The starting months for data. Default is maximum which is 0
    :param m_end: The ending months for data. Default is maximum which is 39
    :return: filtered data as pandas dataframe
    """
    data = get_data()
    data = data[data.prov.isin(prov) & data.con_act.isin(con_act) & data.sex.isin(sex)
                & data.age.isin(ages) & data.measure.isin(measure)]

    return pd.concat([data.iloc[:, :5], data.iloc[:, m_start + 5:m_end + 6]], axis=1, sort=False)


def plot_province(prov="ALL", ages=("ALL",), m_start=0, m_end=39):
    """
    :param prov: String of the province to plot data for
    :param ages: List of the ages to plot data for
    :param m_start: Integer of the start month
    :param m_end: Integer of the end month
    :return: VOID. Just plots the data
    """

    if (len(ages) > 1) and ("ALL" in ages):
        raise Exception("Can't have 'ALL' with other ages")

    # Initialize the data and filter it
    data = filter_data(prov=[prov], ages=ages, m_start=m_start, m_end=m_end)

    data_female = data[data.sex == "F"]
    data_male = data[data.sex == "M"]

    female_total = [0 for i in range(len(CON_ACT))]
    male_total = [0 for i in range(len(CON_ACT))]

    # Get all the sums for the data
    for i in range(len(CON_ACT)):
        for m in MEASURE:
            for a in ages:
                female = data_female[(data_female.con_act == CON_ACT[i]) & (data_female.measure == m) & (data_female.age == a)]
                male = data_male[(data_male.con_act == CON_ACT[i]) & (data_male.measure == m) & (data_male.age == a)]

                if female.shape[0] > 0:
                    female_total[i] += np.sum(np.array(female.iloc[0, 5:].values, dtype=np.int))
                if male.shape[0] > 0:
                    male_total[i] += np.sum(np.array(male.iloc[0, 5:].values, dtype=np.int))

    x = np.arange(len(CON_ACT))  # the label locations
    width = 0.35  # the width of the bars

    # Create the rectangles
    fig, ax = plt.subplots()
    rects2 = ax.bar(x + width / 2, female_total, width, label='Female')
    rects1 = ax.bar(x - width / 2, male_total, width, label='Male')

    # Place the rectangles in the right place
    for rects in (rects1, rects2):
        for rect in rects:
            height = rect.get_height()
            ax.annotate('{}'.format(height),
                        xy=(rect.get_x() + rect.get_width() / 2, height),
                        xytext=(0, 3),  # 3 points vertical offset
                        textcoords="offset points",
                        ha='center', va='bottom')

    # Add some text for labels, title and custom x-axis tick labels, etc.
    ax.set_xticks(x)
    ax.set_xticklabels(CON_ACT)
    ax.legend()
    ax.set_ylabel(f"Number of participants from month {m_start} to month {m_end}")
    ax.set_xlabel("Is the patient using another anti-cancer treatment?")

    if "ALL" in ages:
        str_ages = "18+"
    else:
        str_ages = ", ".join(sorted(ages))

    if "ALL" == prov:
        ax.set_title(f"National data for ages {str_ages}")
    else:
        ax.set_title(f"Data for {prov} for ages {str_ages}")

    fig.tight_layout()
    plt.show()


def _plot_discont(_prov, _ages):
    # Initialize the data and filter it
    data = filter_data(prov=[_prov], ages=_ages)

    data_female = data[(data.sex == 'F')]
    data_male = data[(data.sex == 'M')]

    female_total = np.zeros((len(CON_ACT[1:]), 5))
    male_total = np.zeros((len(CON_ACT[1:]), 5))

    # Get all the sums for the data
    for i in range(len(CON_ACT[1:])):
        for a in _ages:
            female_months = np.zeros(40)
            male_months = np.zeros(40)

            for m in MEASURE:
                female = data_female[(data_female.measure == m) & (data_female.age == a)]
                male = data_male[(data_male.measure == m) & (data_male.age == a)]

                female = female[female.con_act == CON_ACT[i + 1]]
                male = male[male.con_act == CON_ACT[i + 1]]

                if female.shape[0] > 0:
                    female_months += np.array(female.iloc[0, 5:].values, dtype=np.int)
                if male.shape[0] > 0:
                    male_months += np.array(male.iloc[0, 5:].values, dtype=np.int)

                for m in range(0, 40, 10):
                    if female.shape[0] > 0:
                        female_total[i, m // 10] += np.sum(np.array(female.iloc[0, 5 + m: 15 + m].values, dtype=np.int))
                    if male.shape[0] > 0:
                        male_total[i, m // 10] += np.sum(np.array(male.iloc[0, 5 + m: 15 + m].values, dtype=np.int))

            female_rate = 0
            male_rate = 0

            for d in range(1, 40):
                m_before = female_months[d - 1]
                m_after = female_months[d]
                female_rate += np.abs(m_before - m_after)

                if (m_after == 0) or (d == 39):
                    female_rate = female_rate / (d + 1)
                    break

            for d in range(1, 40):
                m_before = male_months[d - 1]
                m_after = male_months[d]
                male_rate += np.abs(m_before - m_after)

                if (m_after == 0) or (d == 39):
                    male_rate = male_rate / (d + 1)
                    break

            female_total[i, 4] = female_rate
            male_total[i, 4] = male_rate

    return female_total, male_total


def plot_discont(prov="ALL", ages=("ALL",)):
    """
    :param prov: String of the province to plot data for
    :param ages: List of the ages to plot data for
    :return: VOID. Just plots the data
    """

    if (len(ages) > 1) and ("ALL" in ages):
        raise Exception("Can't have 'ALL' with other ages")

    totals = _plot_discont(_prov=prov, _ages=ages)

    columns = ('Quarter 1', 'Quarter 2', 'Quarter 3', 'Quarter 4', 'Monthly Rate')
    rows = tuple([("Female " + f"{act}".capitalize()) for act in CON_ACT[1:]] +
                 [("Male " + f"{act}".capitalize()) for act in CON_ACT[1:]])

    # Get some pastel shades for the colors
    colors = [plt.cm.BuPu(np.linspace(0.4, 0.8, len(rows) // 2)), plt.cm.YlOrBr(np.linspace(0.4, 0.8, len(rows) // 2))]
    bar_width = 0.4
    index = [np.arange(len(columns)) + 0.4, np.arange(len(columns))]
    cell_text = [[], []]
    # plt.figure(figsize=(7, 6.5))
    fig, ax = plt.subplots()

    for i in range(len(totals)):
        # Initialize the vertical-offset for the stacked bar chart.
        y_offset = np.zeros(len(columns))

        # Plot bars and create text labels for the table
        data = totals[i]
        height = [0 for i in range(5)]
        width = [0 for i in range(5)]
        x = [0 for i in range(5)]

        for row in range(len(data)):
            bar = plt.bar(index[i], data[row], bar_width, bottom=y_offset, color=colors[i][row])

            for r in range(len(bar)):
                height[r] += bar[r].get_height()
                x[r] = bar[r].get_x()
                width[r] = bar[r].get_width()

            y_offset = y_offset + data[row]
            cell_text[i].append(['%1.1f' % x for x in data[row]])

        for r in range(len(height[:-1])):
            ax.annotate('%1i'%height[r],
                        xy=(x[r] + width[r] / 2, height[r]),
                        xytext=(0, 3),  # 3 points vertical offset
                        textcoords="offset points",
                        ha='center', va='bottom')

    # Add a table at the bottom of the axes
    actual_colors = np.zeros((len(rows), 4))
    actual_colors[:len(rows)//2, :] = colors[0]
    actual_colors[len(rows)//2:, :] = colors[1]

    table = plt.table(cellText=cell_text[0] + cell_text[1],
                          rowLabels=rows,
                          rowColours=actual_colors,
                          colLabels=columns,
                          loc='bottom')

    # Adjust layout to make room for the table:
    table.scale(1, 1.5)
    table.auto_set_font_size(False)
    table.set_fontsize(8)
    plt.subplots_adjust(bottom=0.3)

    plt.xticks([])
    plt.ylabel("Number of participants per quarter")

    if ages[0] == "ALL":
        str_ages = "18+"
    else:
        str_ages = ", ".join(sorted(ages))

    if prov == "ALL":
        plt.title(f'National Quarter Data and Monthly Discontinuation rate for ages {str_ages}')
    else:
        plt.title(f'{prov} Quarter Data and Monthly Discontinuation rate for ages {str_ages}')

    plt.show()


if __name__ == '__main__':
    # Question 1 answer
    plot_province(m_start=8)

    # Question 2 answer
    final = np.zeros(40)
    d = filter_data(prov=("ALL",), con_act=("ALL",), sex=("ALL",), ages=("ALL",))

    final += np.array(d[d.measure == MEASURE[0]].iloc[0, 5:].values, dtype=np.int)
    final += np.array(d[d.measure == MEASURE[1]].iloc[0, 5:].values, dtype=np.int)
    final += np.array(d[d.measure == MEASURE[2]].iloc[0, 5:].values, dtype=np.int)

    disc = 0

    for i in range(1, 40):
        disc += np.abs(final[i - 1] - final[i])

    print(disc/40)
